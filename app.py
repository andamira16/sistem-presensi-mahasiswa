#----------------------------------------- LIBRARY -----------------------------------------#
#===========================================================================================#
from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, Response, send_file, abort
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField
from wtforms.validators import DataRequired, Length
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
import sqlite3
import cv2
import threading
from ultralytics import YOLO
from datetime import datetime, time as dt_time
import os
import time
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from werkzeug.security import generate_password_hash, check_password_hash
import numpy as np
import re
import base64
from ipaddress import ip_address, ip_network
import ipaddress
#===========================================================================================#
#-------------------------------------------------------------------------------------------#





#----------------------------------------- VARIABLE -----------------------------------------#
#============================================================================================#
global_frame = None
detected_objects = []
camera_active = False
detection_active = False
current_model_instance = None
scheduled_start_time = None
scheduled_end_time = None
detected = False
frame_lock = threading.Lock()
#============================================================================================#
#--------------------------------------------------------------------------------------------#



#----------------------------------------- DATABASE -----------------------------------------#
#============================================================================================#
#FLASK
app = Flask(__name__)
app.secret_key = '12345'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///attendance.db' 
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

#SQLAlchemy
db = SQLAlchemy(app)

# Flask-Login Setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Model User untuk Flask-Login
class User(UserMixin):
    def __init__(self, id, username):
        self.id = id
        self.username = username
#CLASS LOGIN
class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=20)])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')

#CLASS REGISTER
class RegisterForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=20)])
    password = PasswordField('Password', validators=[DataRequired(), Length(min=5)])
    submit = SubmitField('Register')

#DIRECTORY DETECTION DARI MODEL
DETECTIONS_DIR = "static/detections"
os.makedirs(DETECTIONS_DIR, exist_ok=True)
#============================================================================================#
#--------------------------------------------------------------------------------------------#



#--------------------------------------- DEF FUNCTION ---------------------------------------#
#============================================================================================#
def init_db():
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()

    # Tabel untuk menyimpan data absensi
    c.execute('''CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        label TEXT,
        waktu TEXT,
        photo_path TEXT,
        matakuliah TEXT,
        kelas TEXT,
        tanggal TEXT
    )''')

    # Tabel untuk menyimpan data pengguna
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        password TEXT
    )''')

    # Tabel baru untuk menyimpan data kehadiran mahasiswa
    c.execute('''CREATE TABLE IF NOT EXISTS submit_attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        submit_label TEXT,
        submit_waktu TEXT,
        submit_photo_path TEXT,
        submit_matakuliah TEXT,
        submit_kelas TEXT,
        submit_tanggal TEXT
    )''')

    conn.commit()
    conn.close()




#FUNCTION SIMPAN HASIL DETEKSI KE DB
def save_to_db(label, waktu, photo_path, matakuliah, kelas, tanggal):
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()
    c.execute('''INSERT INTO attendance (label, waktu, photo_path, matakuliah, kelas, tanggal) 
                 VALUES (?, ?, ?, ?, ?, ?)''', (label, waktu, photo_path, matakuliah, kelas, tanggal))
    conn.commit()
    conn.close()

#FUNCTION SIMPAN DETEKSI FOTO KE DB
def save_detection(label, frame):
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"{label}_{timestamp}.jpg"
    filepath = os.path.join(DETECTIONS_DIR, filename)
    cv2.imwrite(filepath, frame)
    return f"/{filepath}"

def reset_attendance_daily():
    """Menghapus data attendance jika sudah berganti hari"""
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()

    # Ambil tanggal terbaru yang tersimpan di database
    c.execute("SELECT MAX(tanggal) FROM attendance")
    last_date = c.fetchone()[0]

    today = datetime.now().strftime('%Y-%m-%d')

    if last_date and last_date != today:
        print(f"🔄 Menghapus data attendance untuk tanggal lama: {last_date}")
        c.execute("DELETE FROM attendance")
        conn.commit()

    conn.close()

#SET MODEL (KELAS)
@app.route('/set_model', methods=['POST'])
def set_model():
    global current_model_instance, matakuliah, kelas
    data = request.json
    kelas = data.get("kelas")
    matakuliah = data.get("matakuliah")

    if not kelas or not matakuliah:
        return jsonify({"status": "error", "message": "Semua kolom harus diisi."}), 400

    model_mapping = {
        "A": "KelasA.pt",
        "B": "KelasB.pt",
        "C": "KelasC.pt",
        "D": "KelasD.pt",
    }
    model_path = model_mapping.get(kelas)

    if model_path:
        try:
            current_model_instance = YOLO(f"./models/{model_path}")
            print(f"Model {model_path} berhasil disetel.")  # Log jika berhasil
            return jsonify({"status": "success", "message": "Model berhasil disetel."}), 200
        except Exception as e:
            print(f"Error saat memuat model: {e}")  # Log error saat memuat model
            return jsonify({"status": "error", "message": f"Model gagal dimuat: {str(e)}"}), 500
    else:
        return jsonify({"status": "error", "message": "Model tidak ditemukan."}), 400

current_model_instance = None  # Model YOLO akan di-load berdasarkan kelas yang dipilih
matakuliah = ""
kelas = ""

@app.route("/process_frame", methods=["POST"])
def process_frame():
    global detection_active, detected, processed_frame

    if not detection_active:
        print("DEBUG: Deteksi tidak aktif.")
        return jsonify({"success": False, "message": "Deteksi tidak aktif."}), 400

    data = request.get_json()
    image_data = data.get("image")

    if not image_data or not image_data.startswith("data:image/jpeg;base64,"):
        return jsonify({"success": False, "message": "Gambar tidak valid!"}), 400

    try:
        encoded_data = image_data.split(',')[1]
        nparr = np.frombuffer(base64.b64decode(encoded_data), np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        print("DEBUG: Gambar berhasil didekode.")
    except Exception as e:
        print(f"ERROR: Gagal mendecode gambar: {e}")
        return jsonify({"success": False, "message": "Gagal mendecode gambar!"}), 400

    # Proses deteksi menggunakan YOLO
    processed_frame = yolov10_inference(frame)

    if detected:
        # Deteksi selesai, hentikan deteksi
        stop_detection()
        return jsonify({
            "success": True,
            "message": "Wajah terdeteksi!",
            "detected": True
        })

    # Jika tidak ada wajah terdeteksi
    return jsonify({
        "success": True,
        "message": "Tidak ada wajah terdeteksi.",
        "detected": False
    })

@app.route('/start_detection', methods=['POST'])
def start_detection():
    global detection_active, detected
    detection_active = True
    detected = False  # Reset deteksi sebelum memulai
    return jsonify({"status": "success", "message": "Deteksi dimulai!"})

#FUNCTION MODEL
def yolov10_inference(frame, image_size=640, conf_threshold=0.5):
    global current_model_instance, matakuliah, kelas, detected

    if not current_model_instance:
        print("Error: Model belum disetel.")
        return frame

    # Prediksi gambar dengan YOLO
    results = current_model_instance.predict(source=frame, imgsz=image_size, conf=conf_threshold)
    annotated_frame = results[0].plot()

    # Iterasi setiap kotak deteksi yang ditemukan
    for detection in results[0].boxes:
        label_index = int(detection.cls)
        label = current_model_instance.names[label_index]

        # Cek apakah sudah terdeteksi di database HARI INI
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM attendance WHERE label=? AND matakuliah=? AND kelas=? AND tanggal=?", 
                  (label, matakuliah, kelas, datetime.now().strftime('%Y-%m-%d')))
        already_detected_today = c.fetchone()[0] > 0
        conn.close()

        if not already_detected_today:
            # Simpan gambar hasil deteksi
            photo_path = save_detection(label, frame)
            waktu = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            tanggal = datetime.now().strftime('%Y-%m-%d')

            print(f"DEBUG: Deteksi label={label}, kelas={kelas}, matakuliah={matakuliah}, waktu={waktu}, foto_path={photo_path}")
            
            # Simpan data ke database
            save_to_db(label, waktu, photo_path, matakuliah, kelas, tanggal)

            # Tandai bahwa deteksi telah berhasil
            detected = True
            stop_detection()

    return annotated_frame

@app.route('/stop_detection', methods=['POST'])
def stop_detection():
    global detection_active
    detection_active = False
    return jsonify({"status": "success", "message": "Deteksi dihentikan."})

#============================================================================================#
#--------------------------------------------------------------------------------------------#



#----------------------------------------- APP ROUTE -----------------------------------------#
#=============================================================================================#
#LOGIN
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()
        cursor.execute("SELECT id, password FROM users WHERE username = ?", (username,))
        user_data = cursor.fetchone()
        conn.close()

        if user_data:
            user_id, hashed_password = user_data
            if check_password_hash(hashed_password, password):
                user = User(user_id, username)
                login_user(user, remember=True)
                return redirect(url_for('halamandosen'))
            else:
                flash('Password salah.', 'danger')
        else:
            flash('Username tidak ditemukan.', 'danger')

    return render_template('login.html')



@login_manager.user_loader
def load_user(user_id):
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    cursor.execute("SELECT id, username FROM users WHERE id = ?", (user_id,))
    user_data = cursor.fetchone()
    conn.close()

    if user_data:
        return User(id=user_data[0], username=user_data[1])
    return None

#REGISTER
@app.route('/register', methods=['GET', 'POST'])
def register():
    form = RegisterForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data

        conn = sqlite3.connect('attendance.db')
        cursor = conn.cursor()

        # Cek apakah username sudah digunakan
        cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
        if cursor.fetchone():
            flash('Username sudah digunakan. Pilih username lain.', 'danger')
            conn.close()
            return render_template('register.html', form=form)

        # Hash password sebelum disimpan
        hashed_password = generate_password_hash(password)

        # Simpan username dan password yang sudah di-hash
        cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, hashed_password))

        conn.commit()
        conn.close()

        flash('Registrasi berhasil! Silakan login.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html', form=form)

#DASHBOARD
@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

#LOGOUT
@app.route('/logout')
def logout():
    logout_user()
    flash('Anda telah keluar.', 'info')
    return redirect(url_for('login'))

# DATA PRESENSI
@app.route('/attendance_data', methods=['GET'])
def attendance_data():
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()
    
    # Ambil data presensi terbaru
    query = "SELECT label, matakuliah, kelas, waktu, photo_path, tanggal FROM attendance ORDER BY id DESC LIMIT 1"
    data = c.execute(query).fetchone()  
    conn.close()
    
    return render_template('attendance.html', data=data)

@app.route('/delete_latest_attendance', methods=['POST'])
def delete_latest_attendance():
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()

    # Hapus data presensi terbaru
    c.execute("DELETE FROM attendance WHERE id = (SELECT MAX(id) FROM attendance)")
    conn.commit()
    conn.close()

    flash('Data presensi terbaru telah dihapus.', 'success')
    return redirect(url_for('dashboard'))


@app.route('/caripresensi', methods=['GET'])
def caripresensi():
    submit_tanggal = request.args.get('submit_tanggal', '')  # Ambil tanggal dari query parameter
    matakuliah = request.args.get('submit_matakuliah', '')  # Ambil mata kuliah
    kelas = request.args.get('submit_kelas', '')  # Ambil kelas

    if not submit_tanggal:
        return render_template('caripresensi.html', data=[], error="Tanggal harus diisi")

    # Query SQL dengan filter dan urutkan berdasarkan submit_label (nama) secara alfabetis
    query = """
        SELECT id, submit_label, submit_waktu, submit_photo_path
        FROM submit_attendance
        WHERE submit_tanggal = ?
    """
    params = [submit_tanggal]

    if matakuliah:
        query += " AND submit_matakuliah = ?"
        params.append(matakuliah)

    if kelas:
        query += " AND submit_kelas = ?"
        params.append(kelas)

    query += " ORDER BY submit_label ASC"  # Urutkan berdasarkan nama secara alfabetis

    try:
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        c.execute(query, params)
        data = c.fetchall()
        conn.close()
    except Exception as e:
        return render_template('caripresensi.html', data=[], error=f"Terjadi kesalahan: {e}")

    return render_template(
        'caripresensi.html',
        data=data,
        submit_tanggal=submit_tanggal,
        matakuliah=matakuliah,
        kelas=kelas
    )

#DATA PRESENSI
@app.route('/submit_attendance_data', methods=['GET'])
def submit_attendance_data():
    submit_tanggal = request.args.get('submit_tanggal', '')  # Ambil tanggal dari query parameter
    matakuliah = request.args.get('submit_matakuliah', '')  # Ambil mata kuliah
    kelas = request.args.get('submit_kelas', '')  # Ambil kelas

    if not submit_tanggal:
        return render_template('halamandosen.html', data=[], error="Tanggal harus diisi")

    # Query SQL dengan filter dan urutkan berdasarkan submit_label (nama) secara alfabetis
    query = """
        SELECT id, submit_label, submit_waktu, submit_photo_path
        FROM submit_attendance
        WHERE submit_tanggal = ?
    """
    params = [submit_tanggal]

    if matakuliah:
        query += " AND submit_matakuliah = ?"
        params.append(matakuliah)

    if kelas:
        query += " AND submit_kelas = ?"
        params.append(kelas)

    query += " ORDER BY submit_label ASC"  # Urutkan berdasarkan nama secara alfabetis

    try:
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        c.execute(query, params)
        data = c.fetchall()
        conn.close()
    except Exception as e:
        return render_template('halamandosen.html', data=[], error=f"Terjadi kesalahan: {e}")

    return render_template(
        'halamandosen.html',
        data=data,
        submit_tanggal=submit_tanggal,
        matakuliah=matakuliah,
        kelas=kelas
    )


@app.route('/submit_attendance', methods=['POST'])
def submit_attendance():
    if request.method == 'POST':
        label = request.form['label']
        matakuliah = request.form['matakuliah']
        kelas = request.form['kelas']
        waktu = request.form['waktu']
        photo_path = request.form['photo_path']
        tanggal = request.form.get('tanggal', '')
        print(f"Label: {label}, Mata Kuliah: {matakuliah}, Kelas: {kelas}, Waktu: {waktu}, Foto: {photo_path}, Tanggal: {tanggal}")  # Debugging
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        c.execute("INSERT INTO submit_attendance (submit_label, submit_matakuliah, submit_kelas, submit_waktu, submit_photo_path, submit_tanggal) VALUES (?, ?, ?, ?, ?, ?)",
                  (label, matakuliah, kelas, waktu, photo_path, tanggal))
        conn.commit()
        conn.close()

        return redirect(url_for('dashboard'))  # Kembali ke dashboard setelah submit

#HAPUS DATA PRESENSI
@app.route('/delete_attendance', methods=['POST'])
def delete_attendance():
    date = request.form.get('submit_tanggal', '')
    matakuliah = request.form.get('submit_matakuliah', '')
    kelas = request.form.get('submit_kelas', '')
    selected_rows = request.form.get('selected_rows')

    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()

    try:
        if selected_rows:  # Jika ada baris yang dipilih
            selected_ids = [int(x) for x in selected_rows.split(',') if x.isdigit()]
            if selected_ids:
                query = f"DELETE FROM submit_attendance WHERE id IN ({','.join(['?'] * len(selected_ids))})"
                c.execute(query, selected_ids)
                flash(f"{len(selected_ids)} data berhasil dihapus.", "success")
        else:  # Jika tidak ada baris yang dipilih, hapus berdasarkan filter
            query = "DELETE FROM submit_attendance WHERE 1=1"
            params = []

            if date:
                query += " AND submit_tanggal = ?"
                params.append(date)
            if matakuliah:
                query += " AND submit_matakuliah = ?"
                params.append(matakuliah)
            if kelas:
                query += " AND submit_kelas = ?"
                params.append(kelas)

            c.execute(query, params)
            flash("Data presensi berhasil dihapus berdasarkan filter.", "success")

        conn.commit()
    except Exception as e:
        flash(f"Terjadi kesalahan: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for('submit_attendance_data', submit_tanggal=date, submit_matakuliah=matakuliah, submit_kelas=kelas))


@app.route('/export_attendance', methods=['GET'])
def export_attendance():
    # Ambil filter dari query parameters
    date = request.args.get('date', '')  
    matakuliah = request.args.get('matakuliah', '')  
    kelas = request.args.get('kelas', '')

    # Debugging: Print nilai filter untuk memastikan parameternya benar
    print(f"Filter: date={date}, matakuliah={matakuliah}, kelas={kelas}")

    # Buat query SQL dengan filter yang dipilih
    query = "SELECT id, submit_label, submit_waktu, submit_photo_path FROM submit_attendance WHERE 1=1"
    params = []

    if date:
        query += " AND submit_tanggal = ?"
        params.append(date)
    if matakuliah:
        query += " AND submit_matakuliah = ?"
        params.append(matakuliah)
    if kelas:
        query += " AND submit_kelas = ?"
        params.append(kelas)

    # Eksekusi query
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    # Jika tidak ada data yang cocok dengan filter, beri pesan peringatan
    if not rows:
        flash("Tidak ada data untuk diekspor berdasarkan filter yang dipilih.", "warning")
        return redirect(url_for('submit_attendance_data'))

    # Buat file Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Data"
    headers = ['No', 'Nama', 'Waktu', 'Foto']
    ws.append(headers)

    # Format header agar rata tengah
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Menyimpan data ke dalam Excel
    for index, row in enumerate(rows, start=1):
        ws.append([index, row[1], row[2], ""])  # Kolom foto dikosongkan dulu

        # Pastikan path gambar sesuai
        img_path = os.path.abspath(os.path.join("static", "detections", os.path.basename(row[3])))  
        print(f"Checking image path: {img_path}")  # Debugging

        if os.path.exists(img_path):
            print(f"✅ Gambar ditemukan: {img_path}")
            try:
                img = Image(img_path)

                # Tentukan ukuran gambar yang diinginkan
                target_height = 80  
                aspect_ratio = img.width / img.height
                target_width = int(target_height * aspect_ratio)

                img.width = target_width
                img.height = target_height

                # Tambahkan gambar ke dalam cell
                cell_location = f"D{index + 1}"
                ws.add_image(img, cell_location)

                # Sesuaikan tinggi baris dengan tinggi gambar
                ws.row_dimensions[index + 1].height = target_height

                # Atur lebar kolom agar sesuai dengan gambar
                ws.column_dimensions["D"].width = target_width // 10  # Konversi ke ukuran Excel

                # Geser gambar ke tengah cell
                img.anchor = f"D{index + 1}"  # Pastikan gambar ditempatkan dengan benar
            except Exception as e:
                print(f"❌ Error menambahkan gambar: {e}")
                ws[f"D{index + 1}"] = "Error memuat gambar"
        else:
            print(f"⚠️ Gambar tidak ditemukan: {img_path}")
            ws[f"D{index + 1}"] = "Foto tidak ditemukan"

    # Pengaturan lebar kolom otomatis
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Mendapatkan huruf kolom
        for cell in col:
            try:
                if cell.value:  # Jika ada nilai
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(horizontal='center', vertical='center')  # Rata tengah semua teks
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Margin tambahan
        ws.column_dimensions[col_letter].width = adjusted_width

    # Pastikan kolom Foto cukup lebar untuk gambar
    ws.column_dimensions["D"].width = 18  # Lebar tetap untuk kolom Foto

    # Buat nama file berdasarkan filter (tanpa karakter ilegal)
    safe_date = re.sub(r'[^0-9]', '', date) if date else "all"
    safe_matakuliah = re.sub(r'[^a-zA-Z0-9]', '_', matakuliah) if matakuliah else "all"
    safe_kelas = re.sub(r'[^a-zA-Z0-9]', '_', kelas) if kelas else "all"
    filename = f"attendance_{safe_date}_{safe_matakuliah}_{safe_kelas}.xlsx"

    filepath = os.path.join("static", filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True)


@app.route('/halamandosen')
@login_required
def halamandosen():
    return render_template('halamandosen.html')

# Daftar rentang IP jaringan kampus
ALLOWED_IPS = [
    ipaddress.ip_network('192.168.1.2/24'),  # Ganti dengan rentang IP WiFi kampus
    ipaddress.ip_network('192.168.137.1/24')      # Tambahkan rentang lain jika diperlukan
]

def is_ip_allowed(client_ip):
    try:
        ip = ipaddress.ip_address(client_ip)
        return any(ip in network for network in ALLOWED_IPS)
    except ValueError:
        return False

@app.before_request
def restrict_ip():
    # IP klien diperoleh dari header X-Forwarded-For jika menggunakan ngrok
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr).split(',')[0].strip()
    if not is_ip_allowed(client_ip):
        abort(403)  # Forbidden jika IP tidak diizinkan

@app.errorhandler(403)
def forbidden(error):
    return "Access denied. This site can only be accessed using the campus WiFi.", 403

#HALAMAN UTAMA
@app.route('/')
def home():
    return redirect(url_for('dashboard'))
#=============================================================================================#
#---------------------------------------------------------------------------------------------#


#----------------------------------------- RUN APP ------------------------------------------#
#============================================================================================#
if __name__ == '__main__':
    init_db()
    reset_attendance_daily()
    app.run(debug=True)
#============================================================================================#
#--------------------------------------------------------------------------------------------#